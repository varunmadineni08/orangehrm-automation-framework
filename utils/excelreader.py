import openpyxl


def get_login_data():
    final_list=[]
    workbook=openpyxl.load_workbook("utils/Logincreds.xlsx")
    sheet=workbook['loginsheet']
    total_rows=sheet.max_row
    total_cols=sheet.max_column

    for row in range(2,total_rows+1):
        rows_list=[]
        for col in range(1,total_cols+1):
            cell_value=sheet.cell(row,col).value
            rows_list.append(cell_value if cell_value is not None else "")
        final_list.append(rows_list)
    return final_list
